import os
import moviepy.editor as mp
import speech_recognition as sr
from openpyxl import Workbook
import pandas as pd
import re
import openpyxl
import jaconv
from allennlp.models.archival import archive_model, load_archive
from allennlp.predictors.predictor import Predictor
import json
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

def select_video_folder():
    # Tkインスタンスを作成
    root = tk.Tk()
    # Tkインスタンスを非表示にする
    root.withdraw()
    # ユーザーにフォルダ選択を促すメッセージボックスを表示
    messagebox.showinfo("フォルダ選択", "videoフォルダを選択してください。")
    # フォルダ選択ダイアログを表示
    folder_path = filedialog.askdirectory()

    return folder_path

def finish_message():
    # Tkインスタンスを作成
    root = tk.Tk()
    # Tkインスタンスを非表示にする
    root.withdraw()
    # 処理完了のメッセージボックスを表示
    messagebox.showinfo("処理完了", "全ての処理が終了しました。")
    # Tkインスタンスを破棄
    root.destroy()

def convert_video_to_audio(video_path, output_folder):
    """動画から音声を抽出し、一時的な音声ファイルのパスを返す"""
    temp_video_path = os.path.join(output_folder, "temp_video.mp4")
    clip = mp.VideoFileClip(video_path).subclip(0, 7)  # 最初の7秒のみを切り取る
    clip.write_videofile(temp_video_path, codec='libx264', audio_codec='aac')
    
    video_clip = mp.VideoFileClip(temp_video_path)
    audio_clip = video_clip.audio
    temp_audio_path = os.path.join(output_folder, "temp_audio.wav")
    audio_clip.write_audiofile(temp_audio_path)
    
    # リソースを解放
    audio_clip.close()
    video_clip.close()
    clip.close()
    
    os.remove(temp_video_path)  # 一時的なビデオファイルを削除
    return temp_audio_path

def recognize_audio(audio_path, recognizer):
    """音声ファイルを読み込み、テキストに変換する"""
    with sr.AudioFile(audio_path) as source:
        audio = recognizer.record(source)
    return recognizer.recognize_google(audio, language="ja-JP")

def process_videos(video_folder, output_folder):
    """フォルダ内の全動画ファイルを処理し、結果をExcelに保存する"""
    wb = Workbook()
    ws = wb.active
    ws.append(["ラベル", "テキスト", "元のファイル名"])
    recognizer = sr.Recognizer()
    data_rows = []  # 行データを保持するリスト
    
    for filename in os.listdir(video_folder):
        if filename.endswith((".MTS", ".mp4")):
            video_path = os.path.join(video_folder, filename)
            try:
                temp_audio_path = convert_video_to_audio(video_path, output_folder)
                text = recognize_audio(temp_audio_path, recognizer)
                data_rows.append([filename, text, filename])  # 行データをリストに追加
                os.remove(temp_audio_path)  # 一時的な音声ファイルを削除
            except Exception as e:
                print(f"動画 {filename} の処理中にエラーが発生しました: {str(e)}")
                data_rows.append([filename, "エラー", filename])
    
    # ファイル名でリストをソート
    data_rows.sort(key=lambda x: x[0])
    
    # ソートされた順にExcelファイルに行を追加
    for row_data in data_rows:
        ws.append(row_data)
    
    xlsx_save_path = os.path.join(output_folder, "test_text.xlsx")
    wb.save(xlsx_save_path)
    return xlsx_save_path

def clean_excel_data(excel_file_path):
    """Excelファイルのデータをクリーンアップする"""
    df = pd.read_excel(excel_file_path)
    df['ラベル'] = df['ラベル'].apply(lambda x: re.sub(r'\s*\(\d+\)', '', x)).apply(lambda x: re.sub(r'\.\w+$', '', x))
    df.to_excel(excel_file_path, index=False)

    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value:
                cell.value = jaconv.kata2hira(cell.value.replace(' ', ''))
    wb.save(excel_file_path)

def convert_jsonl(excel_path):
    df = pd.read_excel(excel_path)
    output_path = os.path.join(os.path.dirname(excel_path), os.path.splitext(os.path.basename(excel_path))[0] + '_converted.jsonl')
    with open(output_path, 'w', encoding='utf-8') as f:
        for _, row in df.iterrows():
            data = {'label': row[0], 'text': row[1]}
            # json.dumpを使用して辞書をJSON形式で書き込む
            json.dump(data, f, ensure_ascii=False)
            f.write('\n')
    return output_path

def rename_video_file(video_folder, original_filename, prediction_label):
    _, ext = os.path.splitext(original_filename)
    new_filename = f"{prediction_label}{ext}"
    new_filepath = os.path.join(video_folder, new_filename)
    
    counter = 1
    while os.path.exists(new_filepath):
        new_filename = f"{prediction_label}({counter}){ext}"
        new_filepath = os.path.join(video_folder, new_filename)
        counter += 1
    
    original_filepath = os.path.join(video_folder, original_filename)
    os.rename(original_filepath, new_filepath)
    print(f"Renamed '{original_filename}' to '{new_filename}'")

# 現在のスクリプトのディレクトリを取得
current_dir = os.path.dirname(os.path.abspath(__file__))

# メイン処理
video_folder = select_video_folder()
output_folder = current_dir
excel_file_path = process_videos(video_folder, output_folder)
clean_excel_data(excel_file_path)
choice_path = excel_file_path

# ビデオフォルダ内の全ファイル名を取得
video_files = [f for f in os.listdir(video_folder) if f.endswith((".MTS", ".mp4"))]
video_files.sort()  # ファイル名でソート

#　データをjsonに変換し、変換したファイルのパスを取得
converted_jsonl_path = convert_jsonl(choice_path)

# モデルディレクトリのパスを動的に指定
model_dir = os.path.join(current_dir, 'model')

# モデルをアーカイブ
archive_model(model_dir, weights='best.th')

# モデルのアーカイブをロード
archive = load_archive(os.path.join(model_dir, "model.tar.gz"))
predictor = Predictor.from_archive(archive, 'text_classifier')

# テストデータファイルのパスを動的に指定
test_file_path = converted_jsonl_path

# テストデータファイルを読み込み、各テキストに対して予測を実行
with open(test_file_path, 'r', encoding='utf-8') as test_file:
    for i, line in enumerate(test_file):
        json_line = json.loads(line)
        sentence = json_line['text']
        prediction = predictor.predict_json({"sentence": sentence})
        print(f"text: \"{sentence}\", 予測結果: \"{prediction['label']}\"")
        # ファイル名を変更
        rename_video_file(video_folder, video_files[i], prediction['label'])

# 処理が完了した後、作成した.xlsxファイルと.jsonlファイルを削除
os.remove(excel_file_path)
os.remove(converted_jsonl_path)
print(f"処理が完了しました。{excel_file_path} と {converted_jsonl_path} を削除しました。")